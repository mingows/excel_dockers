import json
import os
import datetime as dt # Alias for datetime.datetime
import inspect
import re
import traceback
# from shutil import move # os.rename is closer to fs.renameSync, move is more robust for cross-filesystem
from dateutil import parser as date_parser # For flexible date string parsing (pip install python-dateutil)

# Definir las constantes del módulo
CONSOLE_LOG = True
LOG_LEVEL = ["INFO", "ERROR", "DEBUG"] # Python convention for constants
LOG_FILE_SIZE_LIMIT = 5 * 1024 * 1000  # 5MB
MAX_LOG_FILES = 4

CONSTANTS = {
    "workingPath": os.path.dirname(os.path.abspath(__file__)),
    # "environment": 'DEV',
    # "apiBasePath": 'https://api-manager.scib.dev.corp/api',
    # "clientId": '09390a65-b121-4cd3-bbc4-bebe7a6500d0',
    # "clientSecret": '8c85d438-9d8c-43da-95d5-eaa6da74be65'
}

def get_constants():
    """Returns a JSON string of CONSTANTS."""
    # write_log is called after CONSTANTS is partially initialized.
    write_log(f"Obtenemos las constantes{json.dumps(CONSTANTS)}", "DEBUG", CONSTANTS)
    return json.dumps(CONSTANTS)

def set_constants(global_config_param):
    """Updates the global CONSTANTS dictionary."""
    global CONSTANTS
    CONSTANTS = global_config_param
    write_log(f"Devolvemos las constantes{json.dumps(CONSTANTS)}", "DEBUG", CONSTANTS)
    return CONSTANTS

def rotate_file(log_path, file_name, file_extension):
    """Rotates log files."""
    for i in range(MAX_LOG_FILES - 1, 0, -1):
        current_log = os.path.join(log_path, f"{file_name}{i}.{file_extension}")
        next_log = os.path.join(log_path, f"{file_name}{i + 1}.{file_extension}")
        if os.path.exists(current_log):
            try:
                if os.path.exists(next_log): # os.rename will fail on Windows if next_log exists
                    os.remove(next_log)
                os.rename(current_log, next_log)
            except OSError as e:
                print(f"Error rotating file {current_log} to {next_log}: {e}")

    first_log = os.path.join(log_path, f"{file_name}.{file_extension}")
    second_log = os.path.join(log_path, f"{file_name}1.{file_extension}")
    if os.path.exists(first_log):
        try:
            if os.path.exists(second_log): # os.rename will fail on Windows if second_log exists
                os.remove(second_log)
            os.rename(first_log, second_log)
        except OSError as e:
            print(f"Error rotating file {first_log} to {second_log}: {e}")


def write_log(message, level, global_config):
    """Writes a log message to a file and console, with log rotation."""
    if level in LOG_LEVEL:
        try:
            stack = inspect.stack()
            caller_frame = stack[1]  # The frame that called write_log
            caller_function_name = caller_frame.function
            caller_file_path = caller_frame.filename
            caller_module_name = os.path.basename(caller_file_path)
            
            timestamp = dt.now().isoformat()
            
            if caller_function_name == "<module>": # JS "Object.<anonymous> " equivalent
                log_message_format = f"{timestamp}[{caller_module_name}][{level}] {message}"
            else:
                log_message_format = f"{timestamp}[{caller_module_name}.{caller_function_name}][{level}] {message}"

            if 'workingPath' not in global_config or not global_config['workingPath']:
                print(f"Error: 'workingPath' not configured in global_config for logging. Message: {log_message_format}")
                if CONSOLE_LOG:
                    print(log_message_format)
                return

            log_file_path = os.path.join(global_config['workingPath'], 'logs', 'log.txt')
            log_dir = os.path.dirname(log_file_path)

            if not os.path.exists(log_dir):
                os.makedirs(log_dir, exist_ok=True)

            with open(log_file_path, 'a', encoding='utf-8') as f:
                f.write(log_message_format + '\n')
            
            if os.path.exists(log_file_path) and os.path.getsize(log_file_path) >= LOG_FILE_SIZE_LIMIT:
                rotate_file(log_dir, "log", "txt")
            
            if CONSOLE_LOG:
                print(log_message_format)
        except Exception as e:
            print(f"--- Critical Error in write_log itself ---")
            print(f"Original message: {message}")
            print(f"Logging error: {e}")
            traceback.print_exc()
            print(f"--- End Critical Error in write_log ---")


def write_file(file_path, content, global_config=None):
    """Writes content to a file."""
    try:
        # if global_config:
        #     write_log(f"Escribiendo en el archivo: {file_path}", "DEBUG", global_config)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        # if global_config:
        #     write_log("Escritura finalizada", "DEBUG", global_config)
    except Exception as e:
        log_msg = f"Error writing file {file_path}: {e}"
        if global_config:
            write_log(log_msg, "ERROR", global_config)
        else:
            print(log_msg)

def english_date_validation(date_string):
    """
    Validates if a date string has the format MM/DD/YYYY and is a valid date.
    """
    regex = r"^\d{2}/\d{2}/\d{4}$"
    if not re.match(regex, date_string):
        return False

    parts = date_string.split('/')
    try:
        month = int(parts[0])
        day = int(parts[1])
        year = int(parts[2])

        if not (1 <= month <= 12):
            return False
        # Day check is more accurately handled by dt constructor below
        # if not (1 <= day <= 31):
        #     return False
        if year <= 0: # Or any other year range check
            return False
        
        # dt constructor will raise ValueError for invalid dates like 02/30/2023
        dt(year, month, day) # Python's dt month is 1-indexed
        return True
    except ValueError: # Handles int conversion errors or invalid date for dt constructor
        return False

def append_file(file_path, content, global_config):
    """Appends content to a file."""
    write_log(f"Escribiendo en el archivo: {file_path}", "DEBUG", global_config)
    try:
        with open(file_path, 'a', encoding='utf-8') as f:
            f.write(content)
        write_log("Escritura finalizada", "DEBUG", global_config)
    except Exception as e:
        write_log(f"Error appending to file {file_path}: {e}", "ERROR", global_config)

def format_date(date_input):
    """
    Formats a date (string or datetime object) to YYYY-MM-DD HH:MM:SS.
    Input date_input can be an ISO format string or a datetime object.
    """
    original_date = None
    if isinstance(date_input, str):
        try:
            original_date = date_parser.parse(date_input)
        except ValueError as e:
            raise ValueError(f"Could not parse date string: {date_input}. Error: {e}")
    elif isinstance(date_input, dt):
        original_date = date_input
    else:
        raise TypeError("Input 'date_input' must be a datetime object or a parsable date string.")
    
    return original_date.strftime('%Y-%m-%d %H:%M:%S')

def format_date_xslx(date_obj):
    """
    Converts a datetime object to an Excel serial number.
    Excel epoch is December 30, 1899 (as day 0).
    """
    if not isinstance(date_obj, dt):
        raise TypeError("Input 'date_obj' must be a Python datetime object for format_date_xslx.")

    # Excel epoch: 1899-12-30 00:00:00 UTC (this is day 0 for Excel)
    # Python's datetime month is 1-indexed.
    excel_epoch_dt_utc = dt(1899, 12, 30, tzinfo=dt.now(dt.UTC).tzinfo.utcoffset(None) and dt.UTC or None)


    # Ensure date_obj is treated as UTC if naive, or converted to UTC if aware
    date_obj_utc = None
    if date_obj.tzinfo is None or date_obj.tzinfo.utcoffset(date_obj) is None:
        # Naive datetime, assume it's UTC as per Date.UTC behavior
        date_obj_utc = date_obj.replace(tzinfo=dt.now(dt.UTC).tzinfo.utcoffset(None) and dt.UTC or None)
    else:
        date_obj_utc = date_obj.astimezone(dt.now(dt.UTC).tzinfo.utcoffset(None) and dt.UTC or None)

    # Calculate timestamp in milliseconds since Unix epoch for both dates
    ts_date_obj_ms = date_obj_utc.timestamp() * 1000
    ts_excel_epoch_ms = excel_epoch_dt_utc.timestamp() * 1000
    
    milliseconds_in_a_day = 24 * 60 * 60 * 1000
    
    excel_serial = (ts_date_obj_ms - ts_excel_epoch_ms) / milliseconds_in_a_day
    
    # Excel has a leap year bug for 1900. Dates after 1900-02-28 might need adjustment
    # if the target system strictly follows Excel's bug.
    # This formula is a direct translation of the JS one and common for this conversion.
    # If date_obj_utc is before 1900-03-01 and the system doesn't count 1900-02-29,
    # and the date is >= 1900-03-01, Excel would be +1 compared to a simple day count.
    # However, the provided JS formula is standard.
    # If date_obj_utc > dt(1900, 2, 28, tzinfo=dt.timezone.utc) and excel_serial >= 60:
    #    excel_serial +=1 # This adjustment is sometimes needed.
                          # For direct porting, we keep the JS logic.
    return excel_serial


def set_environment(env=None):
    """Sets environment-specific constants."""
    global CONSTANTS
    environment_key = ""
    if env:
        environment_key = env.lower()
    else:
        while environment_key not in ["dev", "pre", "prod"]:
            environment_key = input('Select an environment (DEV|PRE|PROD):').lower()
    
    # Preserve workingPath
    current_working_path = CONSTANTS.get("workingPath", os.path.dirname(os.path.abspath(__file__)))
    
    config_map = {
        "dev": {
            "environment": 'DEV',
            "apiBasePath": 'https://api-manager.scib.dev.corp/api',
            "clientId": '09390a65-b121-4cd3-bbc4-bebe7a6500d0',
            "clientSecret": '8c85d438-9d8c-43da-95d5-eaa6da74be65'
        },
        "pre": {
            "environment": 'PRE',
            "apiBasePath": 'https://api-manager.scib.pre.corp/api',
            "clientId": '7441a8ed-c07a-4a91-8ad3-d0f70f15c4ba',
            "clientSecret": '17c5b260-865d-4353-97f7-b5d5c2b68a21'
        },
        "prod": {
            "environment": 'PROD',
            "apiBasePath": 'https://api-manager.scib.gs.corp/api',
            "clientId": '1537acf7-c942-4a15-bd76-3dcad80a7fbb',
            "clientSecret": '6eb6c4fc-6eec-4719-9388-c0971edcd558'
        }
    }
    
    chosen_config = config_map.get(environment_key, config_map["dev"]) # Default to dev
    CONSTANTS.update(chosen_config)
    CONSTANTS["workingPath"] = current_working_path # Ensure workingPath is set/preserved
    
    # The JS version returns CONSTANTS, which is fine as it's the modified global.
    return CONSTANTS

def get_environment(env=None):
    """Gets environment-specific constants as a new dictionary."""
    current_working_path = CONSTANTS.get("workingPath", os.path.dirname(os.path.abspath(__file__)))
    # In JS, writeLog(`ENV: ${env}`, "DEBUG2", CONST); DEBUG2 is not in logLevel.
    # Assuming it maps to DEBUG or needs to be added to LOG_LEVEL.
    # write_log(f"ENV: {env}", "DEBUG", CONSTANTS) # Using current global CONSTANTS for logging

    environment_key = ""
    if env:
        environment_key = env.lower()
    else:
        while environment_key not in ["dev", "pre", "prod"]:
            environment_key = input('Select an environment (DEV|PRE|PROD):').lower()

    base_configs = {
        "dev": {
            "environment": 'DEV',
            "apiBasePath": 'https://api-manager.scib.dev.corp/api',
            "clientId": '09390a65-b121-4cd3-bbc4-bebe7a6500d0',
            "clientSecret": '8c85d438-9d8c-43da-95d5-eaa6da74be65'
        },
        "pre": {
            "environment": 'PRE',
            "apiBasePath": 'https://api-manager.scib.pre.corp/api',
            "clientId": '7441a8ed-c07a-4a91-8ad3-d0f70f15c4ba',
            "clientSecret": '17c5b260-865d-4353-97f7-b5d5c2b68a21'
        },
        "prod": {
            "environment": 'PROD',
            "apiBasePath": 'https://api-manager.scib.gs.corp/api',
            "clientId": '1537acf7-c942-4a15-bd76-3dcad80a7fbb',
            "clientSecret": '6eb6c4fc-6eec-4719-9388-c0971edcd558'
        }
    }
    
    # Create a new dictionary for the environment settings
    environment_settings = base_configs.get(environment_key, base_configs["dev"]).copy() # Default to dev
    environment_settings["workingPath"] = current_working_path
    
    return environment_settings

def save_data_with_template_xlsx(data, sheet_name, xlsx_template_path, xlsx_output_path, global_config):
    """
    Placeholder for migrating the saveDataWithTemplateXlsx function from JavaScript.
    The original JS function uses 'xlsx-template'. In Python, you would typically
    use a library like 'openpyxl' to load the template, find placeholders,
    substitute them with values from the 'data' (which corresponds to 'apis' in the JS call),
    and then save the modified workbook.

    Args:
        data: The data to substitute (in JS, it was `{ apis: data }`).
              Here, `data` itself is the object to be used with the key 'apis'.
        sheet_name (str): The name of the sheet in the template to modify.
        xlsx_template_path (str): Path to the .xlsx template file.
        xlsx_output_path (str): Path to save the generated .xlsx file.
        global_config (dict): Global configuration for logging.
    """
    write_log(f"Attempting to process Excel template: {xlsx_template_path} for sheet: {sheet_name}", "INFO", global_config)
    try:
        from openpyxl import load_workbook
        # This is a conceptual implementation. The actual substitution logic
        # depends heavily on how placeholders are defined in your Excel template
        # (e.g., ${apis.some_value}, {{ apis.some_value }}, etc.)
        # and the structure of the `data` object.

        # Assuming `data` is the dictionary that was passed as `apis: data` in JS
        # So, placeholders in Excel might be like ${table:apis.field_name} or similar
        # where 'apis' is the key used in `template.substitute("sheetName", { apis: data })`

        workbook = load_workbook(xlsx_template_path)
        if sheet_name not in workbook.sheetnames:
            write_log(f"Sheet '{sheet_name}' not found in template '{xlsx_template_path}'. Available sheets: {workbook.sheetnames}", "ERROR", global_config)
            raise ValueError(f"Sheet '{sheet_name}' not found in template.")
        
        sheet = workbook[sheet_name]

        # Example: Iterate and replace simple placeholders like "${apis.key}"
        # This is highly simplified. `xlsx-template` likely has more complex substitution.
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # This is a guess at placeholder format. Adjust as needed.
                    # E.g., if placeholder is ${table:CU.date} and data is {"CU": [{"date": "value"}]}
                    # You'd need to parse "CU.date" and access data['CU'][0]['date']
                    # The JS code `template.substitute(sheetName, { apis: data });`
                    # suggests that `data` is the value for a top-level key `apis` in the template context.
                    # So, if a cell has `${apis.some_property}`, you'd look up `data['some_property']`.
                    
                    # Let's assume placeholders are of the form ${key_in_data_object}
                    # and `data` is the dictionary for substitution.
                    # The JS call was `template.substitute(sheetName, { apis: data })`
                    # This means the `data` argument to this Python function corresponds to the `data` in `{apis: data}`.
                    # So, if template has `${apis.value}`, we need to access `data['value']`.
                    
                    # A more robust solution would involve parsing the placeholder string.
                    # For now, this is a conceptual placeholder.
                    # if cell.value.startswith("${") and cell.value.endswith("}"):
                    #    placeholder = cell.value[2:-1]
                    #    # Example: if placeholder is "apis.some_value"
                    #    if placeholder.startswith("apis."):
                    #        actual_key = placeholder.split("apis.", 1)[1]
                    #        if actual_key in data: # data here is the `data` argument to this function
                    #            cell.value = data[actual_key]
                    pass # Actual substitution logic is complex and template-dependent.

        workbook.save(xlsx_output_path)
        write_log(f"Excel file generated and saved to: {xlsx_output_path}", "INFO", global_config)

    except ImportError:
        write_log("openpyxl library is not installed. Please install it: pip install openpyxl", "ERROR", global_config)
        raise
    except Exception as e:
        write_log(f"Error processing Excel template: {e}", "ERROR", global_config)
        traceback.print_exc() # For more detailed error info during development
        raise

def order_settlements_by_month(settlements):
    """
    Sorts a list of settlement dictionaries by the 'settlementMonth' field.
    'settlementMonth' is assumed to be a date string parsable by dateutil.parser.parse.
    """
    if not settlements:
        return []

    def get_sort_key(settlement):
        try:
            month_str = settlement.get('settlementMonth')
            if month_str is None:
                # Handle missing 'settlementMonth' key gracefully for sorting
                return dt.min # Sort items with missing month first
            return date_parser.parse(month_str)
        except (TypeError, ValueError) as e:
            # Handle cases where settlementMonth is present but not a parsable date string
            print(f"Warning: Could not parse settlementMonth '{settlement.get('settlementMonth')}' for sorting: {e}")
            return dt.min # Sort unparsable dates first

    return sorted(settlements, key=get_sort_key)

# Example of how to use (for testing this module itself)
if __name__ == '__main__':
    # Setup a dummy global_config for testing write_log
    test_config = {
        "workingPath": os.path.dirname(os.path.abspath(__file__)),
        "environment": "TEST"
    }
    set_constants(test_config.copy()) # Use a copy

    write_log("Testing log function from utils.py", "INFO", CONSTANTS)
    
    print(f"Initial CONSTANTS: {get_constants()}")
    
    # Test english_date_validation
    print(f"Date '12/31/2023' valid? {english_date_validation('12/31/2023')}") # True
    print(f"Date '02/30/2023' valid? {english_date_validation('02/30/2023')}") # False
    print(f"Date '13/01/2023' valid? {english_date_validation('13/01/2023')}") # False
    print(f"Date 'aa/bb/cccc' valid? {english_date_validation('aa/bb/cccc')}") # False

    # Test format_date
    js_date_str = "2024-11-27T11:55:17.000Z"
    print(f"Formatted date for '{js_date_str}': {format_date(js_date_str)}")
    
    # Test format_date_xslx
    # Note: dt.timezone.utc requires Python 3.2+
    # For older versions, you might need pytz or a fixed offset.
    # For simplicity, assuming Python 3.2+ or that naive datetimes are handled as UTC.
    a_date = dt(2024, 5, 20, 10, 30, 0) # A naive datetime
    print(f"Excel serial for {a_date}: {format_date_xslx(a_date)}")
    a_date_utc = dt(2024, 5, 20, 10, 30, 0, tzinfo=dt.now(dt.UTC).tzinfo.utcoffset(None) and dt.UTC or None) # UTC datetime
    print(f"Excel serial for {a_date_utc} (UTC): {format_date_xslx(a_date_utc)}")


    # Test order_settlements_by_month
    settlements_data = [
        {"settlementMonth": "Mar 2023", "value": 100},
        {"settlementMonth": "Jan 2023", "value": 200},
        {"settlementMonth": "Feb 2023", "value": 150},
        {"settlementMonth": None, "value": 50}, # Test None
        {"settlementMonth": "Invalid Date", "value": 25} # Test invalid
    ]
    ordered_settlements = order_settlements_by_month(settlements_data)
    print("Ordered settlements:")
    for s in ordered_settlements:
        print(s)

    # Test set_environment and get_environment
    # print("Setting environment to 'dev' (interactive if None passed):")
    # dev_config = set_environment() # Will prompt if run interactively
    # print(f"DEV Config from set_environment: {dev_config}")
    # print(f"Global CONSTANTS after set_environment: {CONSTANTS}")

    # print("Getting 'prod' environment settings:")
    # prod_env_settings = get_environment("prod")
    # print(f"PROD Environment settings from get_environment: {prod_env_settings}")
    # print(f"Global CONSTANTS after get_environment (should be unchanged by get_environment): {CONSTANTS}")

    # Placeholder for save_data_with_template_xlsx test
    # Create dummy files and data for a full test
    # dummy_template_path = "dummy_template.xlsx"
    # dummy_output_path = "dummy_output.xlsx"
    # if not os.path.exists(dummy_template_path):
    #     # Create a very simple xlsx for testing if openpyxl is available
    #     try:
    #         from openpyxl import Workbook
    #         wb = Workbook()
    #         sheet = wb.active
    #         sheet.title = "TestDataSheet"
    #         sheet['A1'] = "Hello"
    #         sheet['B1'] = "${apis.world_value}" # Example placeholder
    #         wb.save(dummy_template_path)
    #         print(f"Created dummy template: {dummy_template_path}")
            
    #         dummy_data_for_xlsx = {"world_value": "Python!"}
    #         save_data_with_template_xlsx(dummy_data_for_xlsx, "TestDataSheet", dummy_template_path, dummy_output_path, CONSTANTS)
    #     except ImportError:
    #         print("openpyxl not installed, skipping full save_data_with_template_xlsx test.")
    #     except Exception as e_xlsx:
    #         print(f"Error in save_data_with_template_xlsx test: {e_xlsx}")


