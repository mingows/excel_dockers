import os
import json
import shutil
from openpyxl import load_workbook
# Asumiendo que utils.py está en el mismo directorio o en PYTHONPATH
from utils import write_log, append_file, format_date # order_settlements_by_month, formatDateXslx no se usan directamente aquí pero podrían ser útiles

# Constante global
SAVE_ON = "EXCEL"  # Opciones: "FILE", "EXCEL", "DB"

# Obtener el directorio actual del script (equivalente a __dirname en Node.js)
# __filename en Python es __file__
# __dirname en Python es os.path.dirname(os.path.abspath(__file__))
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

def _perform_substitution(workbook, sheet_name_or_index, data_map, global_config):
    """
    Función auxiliar conceptual para reemplazar marcadores de posición en una hoja de Excel.
    Esta es una simplificación. La lógica real dependerá de la estructura de tus plantillas
    y de cómo xlsx-template maneja los marcadores de posición (ej. ${key.field}, ${table[0].field}).
    """
    try:
        if isinstance(sheet_name_or_index, str):
            if sheet_name_or_index not in workbook.sheetnames:
                write_log(f"Sheet '{sheet_name_or_index}' not found in workbook. Available: {workbook.sheetnames}", "ERROR", global_config)
                return
            sheet = workbook[sheet_name_or_index]
        elif isinstance(sheet_name_or_index, int):
            sheet = workbook.worksheets[sheet_name_or_index]
        else:
            write_log(f"Invalid sheet identifier: {sheet_name_or_index}", "ERROR", global_config)
            return

        # data_map es el objeto que xlsx-template usaría para la sustitución,
        # ej. { "CU": data_line_array } o { "resumeData": resume_data_array }

        # Lógica de sustitución de marcadores de posición (muy simplificada):
        # Deberás adaptar esto a cómo tus plantillas usan los marcadores.
        # Por ejemplo, si tienes ${CU[0].date} y data_map es {"CU": [{"date": "2023-01-01"}]}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Ejemplo básico: reemplazar ${key} con data_map['key']
                    # Esto es solo un ejemplo y no cubre la sintaxis de xlsx-template.
                    # Necesitarás una lógica de parsing de placeholders más robusta.
                    for placeholder_key, placeholder_data in data_map.items():
                        # Esto es demasiado simple para la mayoría de los casos de xlsx-template
                        # que manejan objetos anidados y arrays.
                        # if f"${{{placeholder_key}}}" in cell.value:
                        #    # Esta es una suposición muy básica de formato de placeholder
                        #    # y cómo acceder a los datos.
                        #    # Deberías implementar una lógica de reemplazo más sofisticada.
                        #    pass
                        pass # Implementar lógica de sustitución real aquí
        write_log(f"Conceptual substitution performed for sheet '{sheet.title}'", "DEBUG", global_config)

    except Exception as e:
        write_log(f"Error during conceptual substitution on sheet '{sheet_name_or_index}': {e}", "ERROR", global_config)


def file_save(data, resume_data, file_path, global_config):
    result = {"statusCode": 200, "statusDescription": "OK"}
    try:
        # En JS, data.settlements era la clave. Aquí asumimos que 'data' es el objeto que contiene 'settlements'.
        # Esto necesita adaptarse a la estructura real de 'data' que se pasa a esta función.
        # Si 'data' es directamente la lista de settlements:
        if data and isinstance(data.get("settlements"), list) and len(data["settlements"]) > 0:
            # El path en JS era filePath + "/data#01.json"
            # Usamos os.path.join para construir paths de forma segura
            data_file_path = os.path.join(file_path, "data#01.json")
            append_file(data_file_path, json.dumps(data, indent=2) + "\n", global_config)
            # El log en JS era `Writted data in the file: ${__dirname}` + "/ tmp/data#01.json"
            # Esto parece un error en el log original, ya que filePath es el directorio de datos, no __dirname.
            # Asumiremos que se quería loguear el path real del archivo.
            write_log(f"Written data to file: {data_file_path}", "DEBUG", global_config)
        else:
            result = {"statusCode": 204, "statusDescription": "No Content"}

        # El guardado de resumeData estaba comentado en el JS original.
        # Si se necesita, se puede descomentar y adaptar:
        # resume_file_path = os.path.join(file_path, "resumeData#01.json")
        # append_file(resume_file_path, json.dumps(resume_data, indent=2) + "\n", global_config)
        # write_log(f"Written resume data to file: {resume_file_path}", "DEBUG", global_config)

    except Exception as e:
        write_log(f"Error in file_save: {e}", "ERROR", global_config)
        result["statusCode"] = 400
        result["statusDescription"] = f"Error writing file: {e}"
    return result

def excel_save_resume(resume_data_item, file_path_dir, global_config):
    """
    Guarda una entrada de resumen en un archivo Excel y actualiza la plantilla.
    resume_data_item: Un único diccionario de resumen.
    file_path_dir: Directorio donde se guardará resume.xlsx (ej: 'C:/data/')
    """
    template_path = os.path.join(CURRENT_DIR, "resume_template.xlsx")
    template_tmp_path = os.path.join(CURRENT_DIR, "tmp", "resume_template.xlsx")
    output_path = os.path.join(file_path_dir, "resume.xlsx") # El JS original usaba filePath + "resume.xlsx"

    os.makedirs(os.path.dirname(template_tmp_path), exist_ok=True)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    result = {"statusCode": 200, "statusDescription": "OK"}

    try:
        resume_data_line = [resume_data_item] # xlsx-template espera un array para ${table:...}

        # 1. Crear/Actualizar el archivo de salida (resume.xlsx)
        if not os.path.exists(template_path):
            write_log(f"Template file not found: {template_path}", "ERROR", global_config)
            raise FileNotFoundError(f"Template file not found: {template_path}")

        workbook_output = load_workbook(template_path)
        # Asumimos que la hoja se llama 'resumeData' o es la primera hoja.
        # La sustitución real es compleja y depende de la plantilla.
        _perform_substitution(workbook_output, "resumeData", {"resumeData": resume_data_line}, global_config)
        if os.path.exists(output_path):
            os.remove(output_path)
        workbook_output.save(output_path)

        # 2. Actualizar la plantilla (resume_template.xlsx)
        #    Se añade la nueva línea y luego una línea de marcadores de posición.
        resume_line_tmp_placeholders = {
            "date": "${table:resumeData.date}",
            "tradeDate": "${table:resumeData.tradeDate}",
            "origin": "${table:resumeData.origin}",
            "amount": "${table:resumeData.amount}"
        }
        # Para actualizar la plantilla, cargamos la plantilla actual,
        # le añadimos la nueva línea de datos Y la línea de placeholders.
        # Esto es diferente a cómo xlsx-template podría funcionar (que podría añadir a una tabla existente).
        # Con openpyxl, si la plantilla ya tiene datos, tendríamos que encontrar la última fila
        # y añadir nuevas filas. Por simplicidad aquí, asumimos que la plantilla se regenera
        # con los datos acumulados + placeholders.
        # La lógica original de JS parece sobreescribir la plantilla con la nueva línea + placeholders.

        # Cargar la plantilla original para actualizarla
        workbook_template_update = load_workbook(template_path) # Cargar la plantilla que ya podría tener datos
        
        # Esta parte es conceptual: necesitarías una forma de añadir la nueva línea
        # y luego asegurar que la línea de placeholder sigue al final.
        # Si la plantilla tiene una tabla, tendrías que añadir a esa tabla.
        # Por ahora, vamos a simular la sustitución como si fuera una nueva hoja.
        # Esto NO replica el comportamiento de añadir a una tabla existente.
        
        # Para el TEMPLATE_TMP_PATH y luego TEMPLATE_PATH:
        # La idea es que la plantilla siempre termine con la fila de placeholders.
        # Si la plantilla ya tiene datos, esta lógica es incorrecta.
        # Una mejor aproximación sería leer todas las filas existentes, añadir la nueva,
        # y luego la de placeholders.
        
        # Simplificación: asumimos que la plantilla se regenera con la nueva línea y el placeholder
        all_template_lines = [resume_data_item, resume_line_tmp_placeholders]
        _perform_substitution(workbook_template_update, "resumeData", {"resumeData": all_template_lines}, global_config)

        if os.path.exists(template_tmp_path):
            os.remove(template_tmp_path)
        workbook_template_update.save(template_tmp_path)
        shutil.copyfile(template_tmp_path, template_path)

        write_log(f"Written resume data to file: {output_path}", "DEBUG", global_config)

    except Exception as e:
        now_error_end = format_date(dt.now())
        write_log(f"Error in excel_save_resume: {e}. Hora fin: {now_error_end}", "ERROR", global_config)
        result["statusCode"] = 400
        result["statusDescription"] = f"Error writing file: {e}"
    return result


def excel_save(all_sources_data, file_path_base_dir, global_config):
    """
    Guarda datos de múltiples fuentes en un archivo Excel principal y su resumen.
    all_sources_data: Diccionario donde cada clave es un nombre de fuente (ej. "CU", "NYH")
                      y el valor es un diccionario con {data: [], tmp: [], resume: []}.
    file_path_base_dir: Directorio base (equivalente a __dirname en el script JS original).
    """
    template_data_path = os.path.join(CURRENT_DIR, "Master data_JUL22_template.xlsx")
    # template_data_tmp_path = os.path.join(CURRENT_DIR, "tmp", "Master data_JUL22_template.xlsx") # No se usa en el JS final
    output_data_path = os.path.join(file_path_base_dir, "data", "Master data_JUL22.xlsx")

    template_resume_path = os.path.join(CURRENT_DIR, "resume_template.xlsx")
    # template_resume_tmp_path = os.path.join(CURRENT_DIR, "tmp", "resume_template.xlsx") # No se usa en el JS final
    output_resume_path = os.path.join(file_path_base_dir, "data", "resume.xlsx")

    os.makedirs(os.path.dirname(output_data_path), exist_ok=True)
    # os.makedirs(os.path.dirname(template_data_tmp_path), exist_ok=True) # Si se usara
    os.makedirs(os.path.dirname(output_resume_path), exist_ok=True)
    # os.makedirs(os.path.dirname(template_resume_tmp_path), exist_ok=True) # Si se usara

    result = {"statusCode": 200, "statusDescription": "OK"}

    try:
        if not os.path.exists(template_data_path):
            raise FileNotFoundError(f"Data template not found: {template_data_path}")
        if not os.path.exists(template_resume_path):
            raise FileNotFoundError(f"Resume template not found: {template_resume_path}")

        # Cargar plantillas
        wb_data_output = load_workbook(template_data_path)
        wb_resume_output = load_workbook(template_resume_path)
        
        # Plantillas para actualizar las plantillas originales (copias conceptuales)
        wb_data_template_update = load_workbook(template_data_path)
        wb_resume_template_update = load_workbook(template_resume_path)

        all_resume_lines_for_output = []
        all_resume_lines_for_template_update = []

        for key, source_data in all_sources_data.items():
            sheet_name = key.replace('_', ' ') # ej. Sugar_11 -> Sugar 11
            
            data_lines_for_sheet = source_data['data']    # Array de dicts para la hoja de datos
            tmp_placeholders_for_sheet = source_data['tmp'][0] # Un dict de placeholders
            resume_entry = source_data['resume'][0]       # Un dict para el resumen

            # 1. Sustituir en la plantilla de datos para el archivo de SALIDA
            #    data_map para _perform_substitution sería { sheet_name: data_lines_for_sheet }
            #    Ejemplo: { "CU": [ { "date": "...", "month1": ... }, ... ] }
            _perform_substitution(wb_data_output, sheet_name, {sheet_name: data_lines_for_sheet}, global_config)
            
            # 2. Agregar entrada de resumen para el archivo de SALIDA de resumen
            all_resume_lines_for_output.append(resume_entry)

            # 3. Preparar datos para ACTUALIZAR la PLANTILLA de datos
            #    La plantilla de datos se actualiza con la primera línea de datos reales
            #    seguida de la línea de placeholders.
            data_for_data_template_update = [data_lines_for_sheet[0], tmp_placeholders_for_sheet]
            _perform_substitution(wb_data_template_update, sheet_name, {sheet_name: data_for_data_template_update}, global_config)

            # 4. Agregar entrada de resumen para ACTUALIZAR la PLANTILLA de resumen
            all_resume_lines_for_template_update.append(resume_entry)

        # Finalizar plantillas de resumen
        placeholder_resume_line = {
            "date": "${table:resumeData.date}",
            "tradeDate": "${table:resumeData.tradeDate}",
            "origin": "${table:resumeData.origin}",
            "amount": "${table:resumeData.amount}"
        }
        all_resume_lines_for_template_update.append(placeholder_resume_line)

        _perform_substitution(wb_resume_output, "resumeData", {"resumeData": all_resume_lines_for_output}, global_config)
        _perform_substitution(wb_resume_template_update, "resumeData", {"resumeData": all_resume_lines_for_template_update}, global_config)

        # Guardar archivos de SALIDA
        if os.path.exists(output_data_path): os.remove(output_data_path)
        wb_data_output.save(output_data_path)
        
        if os.path.exists(output_resume_path): os.remove(output_resume_path)
        wb_resume_output.save(output_resume_path)

        # Guardar/Sobrescribir las PLANTILLAS originales
        if os.path.exists(template_data_path): os.remove(template_data_path) # Originalmente se sobreescribía
        wb_data_template_update.save(template_data_path)
        
        if os.path.exists(template_resume_path): os.remove(template_resume_path) # Originalmente se sobreescribía
        wb_resume_template_update.save(template_resume_path)

        write_log(f"Excel files generated/updated in {os.path.dirname(output_data_path)} and templates in {CURRENT_DIR}", "INFO", global_config)

    except Exception as e:
        now_error_end = format_date(dt.now()) if 'dt' in globals() else "Timestamp unavailable"
        write_log(f"Error in excel_save: {e}. Hora fin: {now_error_end}", "ERROR", global_config)
        result["statusCode"] = 400
        result["statusDescription"] = f"Error processing Excel files: {e}"
    return result


def save_data(data, resume_data, file_path, global_config):
    """
    Dispatcher principal para guardar datos.
    'data' y 'resume_data' pueden tener diferentes estructuras dependiendo de SAVE_ON.
    Para SAVE_ON == "EXCEL", se espera que 'data' sea el diccionario `excelData`
    y 'resume_data' no se usa directamente por excel_save, ya que está contenido en `excelData`.
    """
    result = {"statusCode": 200, "statusDescription": "OK"}
    if SAVE_ON == "FILE":
        # Para fileSave, data y resumeData son objetos individuales.
        result = file_save(data, resume_data, file_path, global_config)
    elif SAVE_ON == "EXCEL":
        # Para excelSave, 'data' es el objeto 'excelData' que contiene todo.
        # 'resume_data' no se pasa explícitamente a excel_save, ya que está dentro de 'data'.
        # 'file_path' aquí es el directorio base (__dirname en el script JS original).
        result = excel_save(data, file_path, global_config)
    elif SAVE_ON == "DB":
        write_log(f"Error: SAVE_ON DB is not ready yet: {SAVE_ON}", "ERROR", global_config)
        result["statusCode"] = 400
        result["statusDescription"] = "SAVE_ON DB not implemented"
    else:
        write_log(f"Error: SAVE_ON is not valid: {SAVE_ON}", "ERROR", global_config)
        result["statusCode"] = 400
        result["statusDescription"] = "Invalid SAVE_ON value"
    return result

# Ejemplo de uso (para probar este módulo)
if __name__ == '__main__':
    # Configuración de prueba
    mock_global_config = {
        "workingPath": CURRENT_DIR, # Asegúrate de que utils.py pueda usar esto
        "environment": "TEST_DRIVERS"
    }
    from utils import set_constants
    set_constants(mock_global_config.copy()) # Usar una copia para no modificar el original directamente

    write_log("Testing drivers.py", "INFO", mock_global_config)

    # --- Prueba para file_save ---
    mock_data_for_file = {
        "settlements": [{"tradeDate": "05/29/2025", "origin": "TestOrigin", "settlementMonth": "JUN 25", "settlementAmount": "1.0"}],
        "total": "100"
    }
    mock_resume_for_file = {"date": "2025-05-29...", "tradeDate": "05/29/2025", "origin": "TestOrigin", "amount": 1}
    
    # Crear directorio temporal para file_save si no existe
    test_file_path_dir = os.path.join(CURRENT_DIR, "test_output_drivers_file")
    os.makedirs(test_file_path_dir, exist_ok=True)
    
    # original_save_on = SAVE_ON
    # SAVE_ON = "FILE"
    # print(f"\nTesting file_save (SAVE_ON={SAVE_ON})...")
    # save_data(mock_data_for_file, mock_resume_for_file, test_file_path_dir, mock_global_config)
    # SAVE_ON = original_save_on # Restaurar

    # --- Prueba para excel_save_resume ---
    # Crear plantillas dummy si no existen para la prueba
    dummy_resume_template_path = os.path.join(CURRENT_DIR, "resume_template.xlsx")
    if not os.path.exists(dummy_resume_template_path):
        wb = load_workbook() # Crear nuevo libro
        ws = wb.active
        ws.title = "resumeData"
        ws['A1'] = "Date"
        ws['B1'] = "Trade Date"
        ws['C1'] = "Origin"
        ws['D1'] = "Amount"
        ws['A2'] = "${table:resumeData.date}"
        ws['B2'] = "${table:resumeData.tradeDate}"
        ws['C2'] = "${table:resumeData.origin}"
        ws['D2'] = "${table:resumeData.amount}"
        wb.save(dummy_resume_template_path)
        write_log(f"Created dummy resume template: {dummy_resume_template_path}", "INFO", mock_global_config)

    mock_single_resume_entry = {
        "date": format_date(dt.now()),
        "tradeDate": "05/29/2025",
        "origin": "TestExcelResume",
        "amount": 10
    }
    test_excel_output_dir = os.path.join(CURRENT_DIR, "test_output_drivers_excel", "data") # Simula la estructura de directorios
    os.makedirs(test_excel_output_dir, exist_ok=True)
    
    print(f"\nTesting excel_save_resume...")
    excel_save_resume(mock_single_resume_entry, test_excel_output_dir, mock_global_config)


    # --- Prueba para excel_save (la función más compleja) ---
    dummy_master_template_path = os.path.join(CURRENT_DIR, "Master data_JUL22_template.xlsx")
    if not os.path.exists(dummy_master_template_path):
        wb_master = load_workbook()
        # Crear hojas para CU y NYH como ejemplo
        ws_cu = wb_master.active
        ws_cu.title = "CU"
        ws_cu['A1'] = "Date"; ws_cu['B1'] = "Volume"; ws_cu['C1'] = "Month1";
        ws_cu['A2'] = "${CU[0].date}"; ws_cu['B2'] = "${CU[0].volume}"; ws_cu['C2'] = "${CU[0].month1}";
        
        ws_nyh = wb_master.create_sheet("NYH")
        ws_nyh['A1'] = "Date"; ws_nyh['B1'] = "Volume"; ws_nyh['C1'] = "Month1";
        ws_nyh['A2'] = "${NYH[0].date}"; ws_nyh['B2'] = "${NYH[0].volume}"; ws_nyh['C2'] = "${NYH[0].month1}";
        
        wb_master.save(dummy_master_template_path)
        write_log(f"Created dummy master data template: {dummy_master_template_path}", "INFO", mock_global_config)

    mock_excel_data_all_sources = {
        "CU": {
            "data": [{"date": "29/05/2025", "volume": "1000", "month1": "1.75"}],
            "tmp": [{"date": "${table:CU.date}", "volume": "${table:CU.volume}", "month1": "${table:CU.month1}"}],
            "resume": [{"date": format_date(dt.now()), "tradeDate": "05/29/2025", "origin": "CMEGroup Chicago-CU", "amount": 1}]
        },
        "NYH": {
            "data": [{"date": "29/05/2025", "volume": "500", "month1": "2.05"}],
            "tmp": [{"date": "${table:NYH.date}", "volume": "${table:NYH.volume}", "month1": "${table:NYH.month1}"}],
            "resume": [{"date": format_date(dt.now()), "tradeDate": "05/29/2025", "origin": "CMEGroup New York-NYH", "amount": 1}]
        }
    }
    
    # SAVE_ON = "EXCEL" # Asegurarse de que está en modo EXCEL para saveData
    # print(f"\nTesting excel_save via saveData (SAVE_ON={SAVE_ON})...")
    # # saveData espera que filePath sea el directorio base (__dirname en el script original)
    # save_data(mock_excel_data_all_sources, None, CURRENT_DIR, mock_global_config)

    write_log("Finished testing drivers.py", "INFO", mock_global_config)

