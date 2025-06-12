import json
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Assuming utils.py is in the same directory or accessible in PYTHONPATH
# from utils import write_log, append_file, format_date # format_date from utils.js is for display
# For demonstration, stubs or direct implementations are included if utils.py is not fully aligned.

# --- Stubs/Imports from utils.py (ensure these are correctly in your utils.py) ---
# If your utils.py is set up, these imports would be:
# from utils import write_log, append_file, CONSTANTS (if needed directly)

# Simplified write_log for drivers.py if not importing from a shared utils
def _local_write_log(message, level, global_config, driver_file_name="drivers.py"):
    timestamp = datetime.now().isoformat()
    log_message = f"{timestamp}[{driver_file_name}][{level}] {message}"
    print(log_message)
    if global_config and global_config.get('workingPath'):
        log_dir = os.path.join(global_config['workingPath'], 'logs')
        os.makedirs(log_dir, exist_ok=True)
        try:
            with open(os.path.join(log_dir, 'log.txt'), 'a', encoding='utf-8') as f:
                f.write(log_message + '\n')
        except Exception:
            pass

def _local_append_file(file_path, content, global_config):
    _local_write_log(f"Appending to file: {file_path}", "DEBUG", global_config)
    try:
        with open(file_path, 'a', encoding='utf-8') as f:
            f.write(content)
        _local_write_log("Append finished", "DEBUG", global_config)
    except Exception as e:
        _local_write_log(f"Error appending to file {file_path}: {e}", "ERROR", global_config)

# Use the actual utils.py imports if available and configured
try:
    from utils import write_log, append_file, format_date as util_format_date
except ImportError:
    write_log = _local_write_log
    append_file = _local_append_file
    # format_date from JS utils was for display: YYYY-MM-DD HH:MM:SS
    util_format_date = lambda d_obj: d_obj.strftime('%Y-%m-%d %H:%M:%S') if isinstance(d_obj, datetime) else str(d_obj)


# --- End Stubs/Imports ---

SCRIPT_DIR_DRIVERS = os.path.dirname(os.path.abspath(__file__))
SAVE_ON = "EXCEL"  # Should be FILE, EXCEL, DB

def _substitute_sheet_data(workbook, sheet_name, data_for_sheet, global_config):
    """
    Substitutes data into a specific sheet of a workbook.
    - workbook: openpyxl workbook object.
    - sheet_name: name of the sheet to process.
    - data_for_sheet: A dict like { "KEY_FOR_PLACEHOLDER": [row1_dict, row2_dict, ...] }
                      Example: { "CU": [ {"date": "01/01/2024", "volume": 100}, ... ] }
                      Placeholders in Excel should be like ${KEY_FOR_PLACEHOLDER[0].date} or,
                      if the template implies iteration, ${KEY_FOR_PLACEHOLDER.date} for each row.
                      This implementation assumes placeholders refer to the first row if not a table,
                      or iterates if data is a list.
                      For simplicity, it assumes placeholders are like ${KEY.field} and data is a list of dicts.
    """
    if sheet_name not in workbook.sheetnames:
        write_log(f"Sheet '{sheet_name}' not found in workbook. Available: {workbook.sheetnames}", "ERROR", global_config)
        return

    sheet = workbook[sheet_name]
    
    # This is a simplified substitution. xlsx-template is more powerful.
    # It assumes data_for_sheet has one key, and its value is a list of row data.
    # e.g., data_for_sheet = {"CU": [{"date": "d1", "vol": "v1"}, {"date": "d2", "vol": "v2"}]}
    
    data_key = list(data_for_sheet.keys())[0] # e.g., "CU"
    rows_data = data_for_sheet[data_key]      # e.g., [{"date": "d1", "vol": "v1"}, ...]

    if not isinstance(rows_data, list):
        write_log(f"Data for sheet {sheet_name} key {data_key} is not a list.", "WARN", global_config)
        rows_data = [rows_data] # Treat as a single row if not a list

    # Start writing from the second row (assuming first row is headers)
    # This needs to be more intelligent based on template structure.
    # If template has one placeholder row, we overwrite it. If multiple, we fill them.
    # For now, let's assume we fill starting from row 2.
    
    # A very basic placeholder replacement.
    # This won't handle table expansion like xlsx-template.
    # It will replace placeholders in existing cells.
    
    # If rows_data has more items than placeholder rows in template, this won't insert new rows.
    # If less, it will only fill up to the number of items in rows_data.

    # Let's assume the template has enough rows for the data being passed.
    # For dataLine (1 item) and dataLineTmp (2 items) in JS.
    
    start_row_index = 2 # Default: assume headers in row 1, data starts in row 2

    for row_idx, row_data_dict in enumerate(rows_data):
        current_excel_row = start_row_index + row_idx
        if current_excel_row > sheet.max_row:
            # This basic version doesn't append new rows if data exceeds template rows.
            # write_log(f"Data for sheet {sheet_name} exceeds template rows. Skipping extra data.", "WARN", global_config)
            # A more advanced version would append rows and copy styles.
            # For now, if template has e.g. 1 data row, and we pass 2 rows of data, only 1st is written.
            # The JS code passes dataLine (1 row) or dataLineTmp (2 rows).
            # So, template should have at least 2 data rows with placeholders.
            pass # Let it try to write, openpyxl will handle if row doesn't exist by creating it.

        for col_idx, cell in enumerate(sheet[current_excel_row]): # Iterate cells of the target row
            if cell.value and isinstance(cell.value, str):
                # Attempt to match placeholders like ${CU.date} or ${table:CU.date}
                # This is a simplified match.
                original_cell_value = cell.value
                new_cell_value = original_cell_value
                
                for placeholder_key, val_to_substitute in row_data_dict.items():
                    # Try to match ${data_key.placeholder_key} or ${table:data_key.placeholder_key}
                    # e.g. data_key="CU", placeholder_key="date"
                    # Placeholder in excel: "${CU.date}" or "${table:CU.date}"
                    # More robust: use regex to find all ${...}
                    
                    # Simple replacement:
                    # If cell.value is "${CU.date}" and data_key is "CU", placeholder_key is "date"
                    # then replace with row_data_dict["date"]
                    # This assumes cell.value is *exactly* the placeholder string.
                    
                    # A better way: find all placeholders in cell.value
                    # For now, if cell.value is "${KEY.FIELD}"
                    # e.g. placeholder_in_cell = "CU.date"
                    # e.g. placeholder_in_cell_table = "table:CU.date"
                    
                    # Check for direct match with common patterns
                    target_placeholder1 = f"${{{data_key}.{placeholder_key}}}" # e.g. ${CU.date}
                    target_placeholder2 = f"${{table:{data_key}.{placeholder_key}}}" # e.g. ${table:CU.date}

                    if original_cell_value == target_placeholder1 or original_cell_value == target_placeholder2:
                        new_cell_value = val_to_substitute
                        break # Found and replaced for this cell with this key

                if new_cell_value != original_cell_value:
                     sheet.cell(row=current_excel_row, column=col_idx + 1, value=new_cell_value)


def excel_save(excel_data_map, base_dir_path, global_config):
    # Path definitions (relative to SCRIPT_DIR_DRIVERS, which is where drivers.py is)
    # In JS, __dirname in drivers.js was used.
    template_master_name = "Master data_JUL22_template.xlsx"
    template_resume_name = "resume_template.xlsx"

    template_path = os.path.join(SCRIPT_DIR_DRIVERS, template_master_name)
    template_tmp_path = os.path.join(SCRIPT_DIR_DRIVERS, "tmp", template_master_name)
    output_path = os.path.join(base_dir_path, "data", "Master data_JUL22.xlsx") # base_dir_path is __dirname from index.js

    template_resume_path = os.path.join(SCRIPT_DIR_DRIVERS, template_resume_name)
    template_resume_tmp_path = os.path.join(SCRIPT_DIR_DRIVERS, "tmp", template_resume_name)
    output_resume_path = os.path.join(base_dir_path, "data", "resume.xlsx")

    os.makedirs(os.path.join(SCRIPT_DIR_DRIVERS, "tmp"), exist_ok=True)
    os.makedirs(os.path.join(base_dir_path, "data"), exist_ok=True)

    result = {"statusCode": 200, "statusDescription": "OK"}

    try:
        # Load templates
        try:
            wb_data = load_workbook(template_path)
            wb_resume = load_workbook(template_resume_path)
        except FileNotFoundError as fnf_err:
            write_log(f"Template file not found: {fnf_err}", "ERROR", global_config)
            return {"statusCode": 500, "statusDescription": f"Template file not found: {fnf_err}"}
        except InvalidFileException as ife_err:
            write_log(f"Invalid Excel template file: {ife_err}", "ERROR", global_config)
            return {"statusCode": 500, "statusDescription": f"Invalid Excel template file: {ife_err}"}


        # Create copies for the "tmp" versions (which will become the new templates)
        # This is tricky because openpyxl loads in memory. We save to new tmp path, then load again.
        # Or, more simply, save the modified wb_data to template_tmp_path later.
        # For now, let's work on wb_data and wb_resume for output,
        # and then create wb_data_tmp and wb_resume_tmp for new templates.

        # To create wb_data_tmp, we'd reload the original template
        wb_data_tmp = load_workbook(template_path)
        wb_resume_tmp = load_workbook(template_resume_path)

        all_resume_lines_for_output = []
        all_resume_lines_for_tmp_template = []

        for key, data_item in excel_data_map.items():
            # key is "CU", "NYH", "Sugar_11", etc.
            # data_item is {"data": [...], "tmp": [...], "resume": [...]}
            
            sheet_name_in_excel = key.replace('_', ' ') # e.g., "Sugar_11" -> "Sugar 11"
            write_log(f"Processing sheet: {sheet_name_in_excel} for key: {key}", "DEBUG", global_config)

            # 1. Process data for the output Excel file (wb_data)
            if data_item.get("data"):
                # data_item["data"] is like [{"date": "...", "volume": ..., "month1": ...}]
                data_to_substitute_output = {key: data_item["data"]}
                _substitute_sheet_data(wb_data, sheet_name_in_excel, data_to_substitute_output, global_config)
            
            # 2. Process data for the new template Excel file (wb_data_tmp)
            # This new template should contain the actual data for the first row,
            # and placeholder strings for the second row.
            if data_item.get("data") and data_item.get("tmp"):
                # data_item["data"][0] is the first actual data row
                # data_item["tmp"][0] is the placeholder row object
                combined_data_for_tmp_template = {key: [data_item["data"][0], data_item["tmp"][0]]}
                _substitute_sheet_data(wb_data_tmp, sheet_name_in_excel, combined_data_for_tmp_template, global_config)

            # 3. Collect resume lines
            if data_item.get("resume") and data_item["resume"]:
                all_resume_lines_for_output.append(data_item["resume"][0])
                # For tmp resume, JS copies the actual data, then adds a placeholder marker line
                all_resume_lines_for_tmp_template.append(data_item["resume"][0]) # Add actual

        # Add the placeholder marker line for the tmp resume template
        placeholder_marker_resume_line = {
            "date": "${table:resumeData.date}",
            "tradeDate": "${table:resumeData.tradeDate}",
            "origin": "${table:resumeData.origin}",
            "amount": "${table:resumeData.amount}"
        }
        all_resume_lines_for_tmp_template.append(placeholder_marker_resume_line)

        # Substitute resume data
        if all_resume_lines_for_output:
            _substitute_sheet_data(wb_resume, "resumeData", {"resumeData": all_resume_lines_for_output}, global_config)
        
        if all_resume_lines_for_tmp_template:
            _substitute_sheet_data(wb_resume_tmp, "resumeData", {"resumeData": all_resume_lines_for_tmp_template}, global_config)

        # Save output files
        if os.path.exists(output_path): os.remove(output_path)
        wb_data.save(output_path)
        write_log(f"Saved output data to: {output_path}", "INFO", global_config)

        if os.path.exists(output_resume_path): os.remove(output_resume_path)
        wb_resume.save(output_resume_path)
        write_log(f"Saved output resume to: {output_resume_path}", "INFO", global_config)

        # Save new template files (overwriting old ones)
        if os.path.exists(template_path): os.remove(template_path) # Original template path
        wb_data_tmp.save(template_path) # Save modified tmp as new original template
        write_log(f"Updated template data file at: {template_path}", "INFO", global_config)
        
        # Also save to the explicit tmp path if needed for inspection, though JS overwrites original
        wb_data_tmp.save(template_tmp_path)
        write_log(f"Saved temporary template data to: {template_tmp_path}", "DEBUG", global_config)


        if os.path.exists(template_resume_path): os.remove(template_resume_path) # Original resume template
        wb_resume_tmp.save(template_resume_path) # Save modified tmp as new original resume template
        write_log(f"Updated template resume file at: {template_resume_path}", "INFO", global_config)

        wb_resume_tmp.save(template_resume_tmp_path)
        write_log(f"Saved temporary template resume to: {template_resume_tmp_path}", "DEBUG", global_config)


    except Exception as e:
        error_time = util_format_date(datetime.now())
        write_log(f"Error in excel_save: {e}. Hora fin: {error_time}", "ERROR", global_config)
        import traceback
        write_log(traceback.format_exc(), "ERROR", global_config)
        result["statusCode"] = 500
        result["statusDescription"] = f"Error processing Excel files: {e}"
    
    return result


def file_save(data, resume_data, file_path_dir, global_config):
    # data corresponds to excel_data_map[key]["data"]
    # resume_data corresponds to excel_data_map[key]["resume"]
    # This function in JS seems to be called per data source, but SAVE_ON is global.
    # The JS fileSave appends to data#01.json and resumeData#01.json.
    # This structure is a bit different from how excelSave consumes the aggregated excel_data_map.
    # For now, let's assume data and resume_data are single entries.
    result = {"statusCode": 200, "statusDescription": "OK"}
    try:
        # Assuming 'data' here is the list of settlements (data_item["data"] in excel_save context)
        # and resume_data is the single resume object (data_item["resume"][0])
        
        # The JS `fileSave` expects `data.settlements.length`.
        # If we adapt it to take one item from `excel_data_map`:
        # data_to_save_file = {"settlements": data_item_data, "total": "N/A"} # JS `data` had `total`
        # resume_to_save_file = resume_data_item 
        
        # For simplicity, let's assume this function is called with the full excel_data_map
        # and it iterates or saves the whole thing.
        # However, the original JS `saveData` calls `fileSave` with `data` and `resumeData`
        # which are not clearly defined in that top-level call.
        # Let's stick to the direct port of `excelSave` as it's the one used by `index.js`.
        # If `SAVE_ON = "FILE"` is used, `file_save` needs clearer inputs.

        # Based on JS `fileSave(data, resumeData, filePath, globalConfig)`
        # where `data` has `data.settlements` and `resumeData` is separate.
        # This implies `data` is not the full `excelData` map.
        
        # This function is not directly called by the ported index.py logic if SAVE_ON="EXCEL"
        # So, providing a conceptual port.
        
        data_output_path = os.path.join(file_path_dir, "data#01.json")
        # resume_output_path = os.path.join(file_path_dir, "resumeData#01.json") # JS comments this out

        # Assuming 'data' is a dictionary with a 'settlements' key (list)
        if data.get("settlements") and len(data["settlements"]) > 0:
            append_file(data_output_path, json.dumps(data, indent=2) + "\n", global_config)
            write_log(f"Written data to file: {data_output_path}", "DEBUG", global_config)
        else:
            result["statusCode"] = 204
            result["statusDescription"] = "No Content"
        
        # if resume_data:
        #     append_file(resume_output_path, json.dumps(resume_data, indent=2) + "\n", global_config)
        #     write_log(f"Written resume data to file: {resume_output_path}", "DEBUG", global_config)

    except Exception as e:
        write_log(f"Error in file_save: {e}", "ERROR", global_config)
        result["statusCode"] = 400
        result["statusDescription"] = "Error writing file"
    return result


def save_data(data_map, base_dir_path, global_config):
    """
    Main save function router.
    data_map: The aggregated excelData from index.py.
    base_dir_path: Base directory for output (usually SCRIPT_DIR of index.py).
    """
    result = {"statusCode": 200, "statusDescription": "OK"}
    if SAVE_ON == "FILE":
        # file_save in JS expects a different structure than the aggregated data_map.
        # This part would need refactoring if SAVE_ON="FILE" is to be used with data_map.
        # For now, it's a placeholder.
        write_log("SAVE_ON='FILE' is not fully adapted for aggregated data_map in Python port.", "WARN", global_config)
        # Example: you might iterate data_map and call file_save for each item
        # for key, item_data in data_map.items():
        #    file_save(item_data.get("data_for_file_format"), item_data.get("resume_for_file_format"), ...)
        result["statusCode"] = 501 
        result["statusDescription"] = "SAVE_ON='FILE' not fully implemented for this data structure."
        return result
    elif SAVE_ON == "EXCEL":
        return excel_save(data_map, base_dir_path, global_config)
    elif SAVE_ON == "DB":
        write_log(f"Error: SAVE_ON is not ready yet: {SAVE_ON}", "ERROR", global_config)
        result["statusCode"] = 400
        result["statusDescription"] = "Wrong SAVE_ON value. Expected values: FILE, EXCEL, DB"
        return result
    else:
        write_log(f"Error: SAVE_ON is not valid: {SAVE_ON}", "ERROR", global_config)
        result["statusCode"] = 400
        result["statusDescription"] = "Wrong SAVE_ON value. Expected values: FILE, EXCEL, DB"
        return result

if __name__ == '__main__':
    # Example Usage (for testing drivers.py itself)
    mock_global_config_drivers = {
        "workingPath": SCRIPT_DIR_DRIVERS
    }
    
    # Create dummy template files for testing if they don't exist
    dummy_master_template = os.path.join(SCRIPT_DIR_DRIVERS, "Master data_JUL22_template.xlsx")
    dummy_resume_template = os.path.join(SCRIPT_DIR_DRIVERS, "resume_template.xlsx")

    if not os.path.exists(dummy_master_template):
        from openpyxl import Workbook
        wb_master = Workbook()
        # Add sheets CU, NYH, T2, CORN, RBOB, Sugar 11
        for sheet_title in ["CU", "NYH", "T2", "CORN", "RBOB", "Sugar 11"]:
            ws = wb_master.create_sheet(title=sheet_title)
            ws['A1'] = "Date"
            ws['B1'] = "Volume"
            ws['C1'] = "Month1"
            ws['A2'] = f"${{{sheet_title}.date}}"       # Placeholder for data
            ws['B2'] = f"${{{sheet_title}.volume}}"
            ws['C2'] = f"${{{sheet_title}.month1}}"
            ws['A3'] = f"${{table:{sheet_title}.date}}" # Placeholder for template (tmp)
            ws['B3'] = f"${{table:{sheet_title}.volume}}"
            ws['C3'] = f"${{table:{sheet_title}.month1}}"
        if "Sheet" in wb_master.sheetnames and len(wb_master.sheetnames) > 1 : wb_master.remove(wb_master["Sheet"])
        wb_master.save(dummy_master_template)
        write_log(f"Created dummy master template: {dummy_master_template}", "INFO", mock_global_config_drivers)

    if not os.path.exists(dummy_resume_template):
        from openpyxl import Workbook
        wb_resume_tpl = Workbook()
        ws_resume = wb_resume_tpl.active
        ws_resume.title = "resumeData"
        ws_resume['A1'] = "Date"
        ws_resume['B1'] = "TradeDate"
        ws_resume['C1'] = "Origin"
        ws_resume['D1'] = "Amount"
        ws_resume['A2'] = "${resumeData.date}"
        ws_resume['B2'] = "${resumeData.tradeDate}"
        ws_resume['C2'] = "${resumeData.origin}"
        ws_resume['D2'] = "${resumeData.amount}"
        # Add a second row for the tmp part
        ws_resume['A3'] = "${table:resumeData.date}"
        ws_resume['B3'] = "${table:resumeData.tradeDate}"
        ws_resume['C3'] = "${table:resumeData.origin}"
        ws_resume['D3'] = "${table:resumeData.amount}"
        wb_resume_tpl.save(dummy_resume_template)
        write_log(f"Created dummy resume template: {dummy_resume_template}", "INFO", mock_global_config_drivers)

    # Mock excel_data_map similar to what index.py would produce
    mock_excel_data = {
        "CU": {
            "data": [{"date": "29/05/2024", "volume": 1000.0, "month1": 170.0}],
            "tmp": [{"date": "${table:CU.date}", "volume": "${table:CU.volume}", "month1": "${table:CU.month1}"}],
            "resume": [{"date": "2024-05-30 10:00:00", "tradeDate": "29/05/2024", "origin": "CMEGroup Chicago-CU", "amount": 1}]
        },
        "NYH": {
            "data": [{"date": "29/05/2024", "volume": 2000.0, "month1": 180.0}],
            "tmp": [{"date": "${table:NYH.date}", "volume": "${table:NYH.volume}", "month1": "${table:NYH.month1}"}],
            "resume": [{"date": "2024-05-30 10:01:00", "tradeDate": "29/05/2024", "origin": "CMEGroup NewYork-NYH", "amount": 1}]
        }
    }
    
    write_log("Testing excel_save...", "INFO", mock_global_config_drivers)
    # base_dir_path for testing drivers.py will be SCRIPT_DIR_DRIVERS
    # In real use, index.py passes its own SCRIPT_DIR
    result_save = excel_save(mock_excel_data, SCRIPT_DIR_DRIVERS, mock_global_config_drivers)
    write_log(f"excel_save test result: {result_save}", "INFO", mock_global_config_drivers)

